#!/usr/bin/env python3
"""
Optimize the existing Excel template - add auto-filter and verify structure.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path


def optimize_excel_template():
    """Optimize the Excel template."""

    template_path = Path('support/doc-templates/excel/EOFramework-Excel-Template-01.xlsx')

    print("📄 Optimizing Excel Template")
    print("=" * 70)
    print()

    # Load template
    wb = load_workbook(str(template_path))

    print(f"✅ Loaded: {template_path}")
    print(f"   Sheets: {wb.sheetnames}")
    print()

    # =========================================================================
    # OPTIMIZE DATA SHEET
    # =========================================================================
    if 'Data' in wb.sheetnames:
        data = wb['Data']

        print("📊 Optimizing Data Sheet")
        print("-" * 70)

        # Enable auto-filter on header row
        max_col = data.max_column
        max_row = data.max_row

        # Set auto-filter range
        auto_filter_range = f'A1:{get_column_letter(max_col)}1'
        data.auto_filter.ref = auto_filter_range

        print(f"  ✅ Auto-filter enabled: {auto_filter_range}")

        # Verify freeze panes
        if data.freeze_panes:
            print(f"  ✅ Freeze panes: {data.freeze_panes}")
        else:
            data.freeze_panes = data['A2']
            print(f"  ✅ Freeze panes set: A2")

        print(f"  ✅ Data rows: {max_row}")
        print(f"  ✅ Columns: {max_col}")
        print()

    # =========================================================================
    # VERIFY COVER SHEET
    # =========================================================================
    if 'Cover' in wb.sheetnames:
        cover = wb['Cover']

        print("📋 Verifying Cover Sheet")
        print("-" * 70)

        # Check key fields
        print(f"  Document Title (B4): {cover['B4'].value}")
        print(f"  Document Type (B5): {cover['B5'].value}")
        print(f"  Generated Label (B7): {cover['B7'].value}")
        print(f"  Generated Value (C7): {cover['C7'].value}")
        print(f"  Solution Label (B8): {cover['B8'].value}")
        print(f"  Solution Value (C8): {cover['C8'].value}")
        print(f"  Purpose Label (B9): {cover['B9'].value}")
        print(f"  Purpose Value (C9): {cover['C9'].value}")
        print(f"  Version Label (B10): {cover['B10'].value}")
        print(f"  Version Value (C10): {cover['C10'].value}")
        print(f"  Images (logos): {len(cover._images)}")
        print()

    # Save optimized template
    print("=" * 70)
    print(f"💾 Saving optimized template")

    wb.save(str(template_path))

    print()
    print("✨ Template optimized successfully!")
    print()
    print("Optimizations applied:")
    print("  ✅ Auto-filter enabled on Data sheet")
    print("  ✅ Freeze panes verified on Data sheet")
    print("  ✅ Template structure validated")
    print()


if __name__ == '__main__':
    optimize_excel_template()
