#!/usr/bin/env python3
"""
Create Excel template with cover sheet and styled data template.
Similar to PowerPoint template approach.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime


def create_excel_template():
    """Create Excel template with cover and data sheets."""

    output_file = Path('support/doc-templates/excel/EOFramework-Excel-Template-01.xlsx')
    logo_path = Path('support/doc-templates/powerpoint')

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    print("📄 Creating Excel Template")
    print("=" * 70)

    # =========================================================================
    # SHEET 1: COVER PAGE
    # =========================================================================
    print("\n📋 Sheet 1: Cover Page")
    print("-" * 70)

    cover = wb.create_sheet("Cover", 0)

    # Set column widths for cover layout
    cover.column_dimensions['A'].width = 5
    cover.column_dimensions['B'].width = 30
    cover.column_dimensions['C'].width = 30
    cover.column_dimensions['D'].width = 5

    # Title styling
    title_font = Font(name='Calibri', size=24, bold=True, color='1F4E78')
    subtitle_font = Font(name='Calibri', size=14, color='44546A')
    label_font = Font(name='Calibri', size=11, bold=True, color='44546A')
    value_font = Font(name='Calibri', size=11, color='44546A')

    # Row 2-3: Spacing
    cover.row_dimensions[1].height = 20
    cover.row_dimensions[2].height = 5

    # Row 4: Document Title (placeholder)
    cover['B4'] = '[DOCUMENT TITLE]'
    cover['B4'].font = title_font
    cover['B4'].alignment = Alignment(horizontal='left', vertical='center')
    cover.merge_cells('B4:C4')
    cover.row_dimensions[4].height = 30

    # Row 5: Document Type (placeholder)
    cover['B5'] = '[Document Type]'
    cover['B5'].font = subtitle_font
    cover['B5'].alignment = Alignment(horizontal='left', vertical='center')
    cover.merge_cells('B5:C5')
    cover.row_dimensions[5].height = 20

    # Row 6-7: Spacing
    cover.row_dimensions[6].height = 10
    cover.row_dimensions[7].height = 5

    # Row 8: Client Logo placeholder
    cover['B8'] = '[Client Logo]'
    cover['B8'].font = Font(name='Calibri', size=10, italic=True, color='7F7F7F')
    cover['B8'].alignment = Alignment(horizontal='center', vertical='center')
    cover['B8'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    cover.merge_cells('B8:B10')
    cover.row_dimensions[8].height = 60

    # Insert client logo if available
    client_logo = logo_path / 'client_logo.png'
    if client_logo.exists():
        img = XLImage(str(client_logo))
        img.width = 150
        img.height = 60
        cover.add_image(img, 'B8')
        print("  ✅ Client logo placeholder added")
    else:
        print("  ⚠️  Client logo not found, using text placeholder")

    # Row 11: Spacing
    cover.row_dimensions[11].height = 10

    # Row 12: Consulting Company Logo placeholder
    cover['B12'] = '[Consulting Company Logo]'
    cover['B12'].font = Font(name='Calibri', size=10, italic=True, color='7F7F7F')
    cover['B12'].alignment = Alignment(horizontal='center', vertical='center')
    cover['B12'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    cover.merge_cells('B12:B14')
    cover.row_dimensions[12].height = 60

    # Insert consulting logo if available
    consulting_logo = logo_path / 'consulting_company_logo.png'
    if consulting_logo.exists():
        img = XLImage(str(consulting_logo))
        img.width = 150
        img.height = 60
        cover.add_image(img, 'B12')
        print("  ✅ Consulting company logo placeholder added")
    else:
        print("  ⚠️  Consulting company logo not found, using text placeholder")

    # Row 15: Spacing
    cover.row_dimensions[15].height = 10

    # Row 16: EO Framework Logo placeholder
    cover['B16'] = '[EO Framework Logo]'
    cover['B16'].font = Font(name='Calibri', size=10, italic=True, color='7F7F7F')
    cover['B16'].alignment = Alignment(horizontal='center', vertical='center')
    cover['B16'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    cover.merge_cells('B16:B18')
    cover.row_dimensions[16].height = 60

    # Insert EO Framework logo if available
    eo_logo = logo_path / 'eo-framework-logo-real.png'
    if eo_logo.exists():
        img = XLImage(str(eo_logo))
        img.width = 200
        img.height = 60
        cover.add_image(img, 'B16')
        print("  ✅ EO Framework logo placeholder added")
    else:
        print("  ⚠️  EO Framework logo not found, using text placeholder")

    # Row 19-20: Spacing
    cover.row_dimensions[19].height = 15
    cover.row_dimensions[20].height = 5

    # Row 21: Metadata - Generated Date
    cover['B21'] = 'Generated:'
    cover['B21'].font = label_font
    cover['C21'] = f'{datetime.now().strftime("%B %d, %Y")}'
    cover['C21'].font = value_font

    # Row 22: Metadata - Solution
    cover['B22'] = 'Solution:'
    cover['B22'].font = label_font
    cover['C22'] = '[Solution Name]'
    cover['C22'].font = value_font

    # Row 23: Metadata - Purpose
    cover['B23'] = 'Purpose:'
    cover['B23'].font = label_font
    cover['C23'] = '[Document Purpose]'
    cover['C23'].font = value_font

    # Row 24: Metadata - Version
    cover['B24'] = 'Version:'
    cover['B24'].font = label_font
    cover['C24'] = '1.0'
    cover['C24'].font = value_font

    print("  ✅ Cover page layout complete")
    print("  ✅ Metadata fields added")

    # =========================================================================
    # SHEET 2: DATA TEMPLATE
    # =========================================================================
    print("\n📊 Sheet 2: Data Template")
    print("-" * 70)

    data = wb.create_sheet("Data", 1)

    # Sample header row with styling
    headers = ['Column 1', 'Column 2', 'Column 3', 'Column 4', 'Column 5']

    # Header styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set headers
    for col_idx, header in enumerate(headers, 1):
        cell = data.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    print("  ✅ Header row configured")

    # Sample data rows with alternating colors
    sample_data = [
        ['Sample Data 1', 'Value A', '100', 'Active', 'Note 1'],
        ['Sample Data 2', 'Value B', '200', 'Pending', 'Note 2'],
        ['Sample Data 3', 'Value C', '300', 'Complete', 'Note 3'],
        ['Sample Data 4', 'Value D', '400', 'Active', 'Note 4'],
    ]

    # Alternate row fill
    even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Data styling
    data_font = Font(name='Calibri', size=11, color='000000')
    data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Add sample data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = data.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = even_fill
            else:
                cell.fill = odd_fill

    print("  ✅ Sample data rows with alternating colors")

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        data.column_dimensions[col_letter].width = 20

    print("  ✅ Column widths configured")

    # Freeze header row
    data.freeze_panes = data['A2']
    print("  ✅ Freeze panes enabled")

    # Enable auto-filter
    data.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    print("  ✅ Auto-filter enabled")

    # Save template
    print("\n" + "=" * 70)
    print(f"💾 Saving template: {output_file}")

    # Ensure directory exists
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(output_file))

    print()
    print("✨ Excel template created successfully!")
    print(f"   Location: {output_file.absolute()}")
    print()
    print("Template structure:")
    print("  📋 Sheet 1: Cover - Professional cover page with logos and metadata")
    print("  📊 Sheet 2: Data - Styled template with headers and alternating rows")
    print()
    print("Features:")
    print("  ✅ 3 logo placeholders (Client, Consulting, EO Framework)")
    print("  ✅ Metadata fields (Generated, Solution, Purpose, Version)")
    print("  ✅ Styled header row (blue background, white text, bold)")
    print("  ✅ Alternating row colors (gray/white)")
    print("  ✅ Borders on all cells")
    print("  ✅ Auto-filter enabled")
    print("  ✅ Freeze panes on header row")
    print("  ✅ Professional fonts (Calibri)")


if __name__ == '__main__':
    create_excel_template()
