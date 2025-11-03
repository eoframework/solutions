#!/usr/bin/env python3
"""
Create a clean Excel template without Table objects that cause conflicts.
Uses simple ranges with AutoFilter instead of Excel Tables.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime


def create_clean_excel_template():
    """Create clean Excel template avoiding Table/AutoFilter conflicts."""

    output_file = Path('support/doc-templates/excel/EOFramework-Excel-Template-01.xlsx')
    logo_path = Path('support/doc-templates/assets/logos')

    # Create new workbook (not loading any existing template)
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    print("📄 Creating Clean Excel Template")
    print("=" * 70)

    # =========================================================================
    # SHEET 1: COVER PAGE
    # =========================================================================
    print("\n📋 Sheet 1: Cover Page")
    print("-" * 70)

    cover = wb.create_sheet("Cover", 0)

    # Set column widths
    cover.column_dimensions['A'].width = 5
    cover.column_dimensions['B'].width = 30
    cover.column_dimensions['C'].width = 13
    cover.column_dimensions['D'].width = 5

    # Title styling
    title_font = Font(name='Calibri', size=24, bold=True, color='1F4E78')
    subtitle_font = Font(name='Calibri', size=14, color='44546A')
    label_font = Font(name='Calibri', size=11, bold=True, color='44546A')
    value_font = Font(name='Calibri', size=11, color='44546A')

    # Row 1-2: Top spacing
    cover.row_dimensions[1].height = 20
    cover.row_dimensions[2].height = 10

    # Row 3: Client Logo
    cover.row_dimensions[3].height = 60
    client_logo = logo_path / 'client_logo.png'
    if client_logo.exists():
        img = XLImage(str(client_logo))
        img.width = 150
        img.height = 50
        img.anchor = 'B3'
        cover.add_image(img)
        print("  ✅ Client logo added at B3")

    # Row 4: Document Title
    cover['B4'] = '[DOCUMENT TITLE]'
    cover['B4'].font = title_font
    cover['B4'].alignment = Alignment(horizontal='left', vertical='center')
    cover.row_dimensions[4].height = 30

    # Row 5: Document Type
    cover['B5'] = '[Document Type]'
    cover['B5'].font = subtitle_font
    cover['B5'].alignment = Alignment(horizontal='left', vertical='center')
    cover.row_dimensions[5].height = 20

    # Row 6: Spacing
    cover.row_dimensions[6].height = 20

    # Rows 7-10: Metadata
    cover['B7'] = 'Generated:'
    cover['B7'].font = label_font
    cover['C7'] = datetime.now().strftime('%B %d, %Y')
    cover['C7'].font = value_font

    cover['B8'] = 'Solution:'
    cover['B8'].font = label_font
    cover['C8'] = '[Solution Name]'
    cover['C8'].font = value_font

    cover['B9'] = 'Purpose:'
    cover['B9'].font = label_font
    cover['C9'] = '[Document Purpose]'
    cover['C9'].font = value_font

    cover['B10'] = 'Version:'
    cover['B10'].font = label_font
    cover['C10'] = '1.0'
    cover['C10'].font = value_font

    # Row 11-12: Spacing
    cover.row_dimensions[11].height = 20
    cover.row_dimensions[12].height = 10

    # Row 13: Consulting Company Logo
    cover.row_dimensions[13].height = 60
    consulting_logo = logo_path / 'consulting_company_logo.png'
    if consulting_logo.exists():
        img = XLImage(str(consulting_logo))
        img.width = 150
        img.height = 50
        img.anchor = 'B13'
        cover.add_image(img)
        print("  ✅ Consulting company logo added at B13")

    # Row 14: Spacing
    cover.row_dimensions[14].height = 10

    # Row 15: EO Framework Logo
    cover.row_dimensions[15].height = 60
    eo_logo = logo_path / 'eo-framework-logo-real.png'
    if eo_logo.exists():
        img = XLImage(str(eo_logo))
        img.width = 200
        img.height = 50
        img.anchor = 'B15'
        cover.add_image(img)
        print("  ✅ EO Framework logo added at B15")

    print("  ✅ Cover page layout complete")

    # =========================================================================
    # SHEET 2: DATA TEMPLATE
    # =========================================================================
    print("\n📊 Sheet 2: Data Template")
    print("-" * 70)

    data = wb.create_sheet("Data", 1)

    # Sample headers
    headers = ['Column 1', 'Column 2', 'Column 3', 'Column 4', 'Column 5']

    # Styling
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')  # Size 12
    header_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # Gray header
    header_alignment = Alignment(horizontal='center', vertical='center')

    data_font = Font(name='Calibri', size=12, color='000000')  # Size 12
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)  # Center vertical + indent for padding

    # Alternating row colors for better readability
    even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light gray
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')   # White

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = data.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    print("  ✅ Header row created")

    # Sample data (just a few rows to show styling)
    sample_data = [
        ['Sample Data 1', 'Value A', '100', 'Active', 'Note 1'],
        ['Sample Data 2', 'Value B', '200', 'Pending', 'Note 2'],
        ['Sample Data 3', 'Value C', '300', 'Complete', 'Note 3'],
        ['Sample Data 4', 'Value D', '400', 'Active', 'Note 4'],
    ]

    # Write sample data with alternating colors
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = data.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = even_fill
            else:
                cell.fill = odd_fill

    print("  ✅ Sample data rows added")

    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        data.column_dimensions[get_column_letter(col_idx)].width = 20

    print("  ✅ Column widths set")

    # Set minimum row heights with auto-adjust capability (maintains padding while allowing content expansion)
    # Note: Setting explicit height, but wrap_text + vertical center provides the padding effect
    data.row_dimensions[1].height = 32  # Header row - minimum height with padding
    for row_idx in range(2, len(sample_data) + 2):
        data.row_dimensions[row_idx].height = 26  # Data rows - minimum height, will auto-expand with content

    print("  ✅ Row heights set (auto-adjusts to content with padding)")

    # Freeze header row
    data.freeze_panes = 'A2'
    print("  ✅ Freeze panes enabled at A2")

    # Enable auto-filter (without using Table object)
    data.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    print("  ✅ Auto-filter enabled (range-based, not Table)")

    # Save template
    print("\n" + "=" * 70)
    print(f"💾 Saving clean template: {output_file}")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_file))

    print()
    print("✨ Clean Excel template created successfully!")
    print()
    print("Template structure:")
    print("  📋 Sheet 1: Cover - 3 logos, metadata fields")
    print("  📊 Sheet 2: Data - Styled headers, alternating rows, auto-filter")
    print()
    print("Key fixes:")
    print("  ✅ No Excel Table objects (avoids Table/AutoFilter conflict)")
    print("  ✅ Simple range-based AutoFilter")
    print("  ✅ Clean image anchoring (avoids drawing corruption)")
    print("  ✅ Proper styling without Table formatting")
    print()


if __name__ == '__main__':
    create_clean_excel_template()
